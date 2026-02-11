"""
Sistema de Gestión de Reportes Excel - Aplicación Web Flask
Autor: Sistema Integrado de Reportes
Versión: 2.0
Fecha: 2024
"""

import os
import sys
import logging
from datetime import datetime, timedelta
from flask import Flask, render_template, request, send_file, jsonify, redirect, url_for, flash, session
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import secrets
from functools import wraps

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Inicializar Flask
app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Carpeta para almacenar archivos generados
UPLOAD_FOLDER = 'generated_reports'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Decorador para requerir autenticación (simulada)
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Por favor inicie sesión para acceder a esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Página de inicio
@app.route('/')
def index():
    """Página principal del sistema"""
    return render_template('index.html', 
                         titulo="Sistema de Reportes Excel",
                         usuario=session.get('user', 'Invitado'))

# Página de login (simulado)
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de inicio de sesión"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Autenticación simulada
        if username == 'admin' and password == 'admin123':
            session['user'] = username
            flash('¡Bienvenido al sistema!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Credenciales incorrectas. Intente con admin/admin123', 'danger')
    
    return render_template('login.html')

# Página de logout
@app.route('/logout')
def logout():
    """Cerrar sesión"""
    session.pop('user', None)
    flash('Sesión cerrada exitosamente.', 'info')
    return redirect(url_for('index'))

# Dashboard principal
@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard principal del sistema"""
    # Contar archivos generados
    reportes = []
    if os.path.exists(UPLOAD_FOLDER):
        archivos = os.listdir(UPLOAD_FOLDER)
        for archivo in archivos:
            if archivo.endswith('.xlsx'):
                ruta = os.path.join(UPLOAD_FOLDER, archivo)
                tamaño = os.path.getsize(ruta)
                fecha = datetime.fromtimestamp(os.path.getctime(ruta))
                reportes.append({
                    'nombre': archivo,
                    'tamaño': tamaño,
                    'fecha': fecha.strftime('%Y-%m-%d %H:%M:%S'),
                    'ruta': ruta
                })
        # Ordenar por fecha (más recientes primero)
        reportes.sort(key=lambda x: x['fecha'], reverse=True)
    
    # Estadísticas
    estadisticas = {
        'total_reportes': len(reportes),
        'total_mapa_flujo': len([r for r in reportes if 'Mapa_Flujo' in r['nombre']]),
        'total_planificacion': len([r for r in reportes if 'Planificacion' in r['nombre']]),
        'total_foda': len([r for r in reportes if 'FODA' in r['nombre']]),
        'total_completo': len([r for r in reportes if 'Completo' in r['nombre']]),
        'tamaño_total': sum(r['tamaño'] for r in reportes)
    }
    
    return render_template('dashboard.html',
                         usuario=session.get('user'),
                         reportes=reportes[:10],  # Solo los 10 más recientes
                         estadisticas=estadisticas)

# Página de generación de reportes
@app.route('/generar')
@login_required
def generar():
    """Página para generar nuevos reportes"""
    return render_template('generar.html', usuario=session.get('user'))

# Función para generar Mapa de Flujo
def generar_mapa_flujo():
    """Genera el reporte de Mapa de Flujo"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard Resumen"
        
        # Estilos
        dark_bg = PatternFill(start_color="0A0F2B", end_color="0A0F2B", fill_type="solid")
        header_fill = PatternFill(start_color="151F4D", end_color="151F4D", fill_type="solid")
        card_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        success_fill = PatternFill(start_color="4CAF50", end_color="2E7D32", fill_type="solid")
        warning_fill = PatternFill(start_color="FF9800", end_color="F57C00", fill_type="solid")
        info_fill = PatternFill(start_color="2196F3", end_color="1976D2", fill_type="solid")
        
        title_font = Font(color="FFFFFF", bold=True, size=16)
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Título
        ws.merge_cells('A1:N1')
        ws['A1'] = "📊 MAPA DE FLUJO DE ACTIVIDADES"
        ws['A1'].font = title_font
        ws['A1'].fill = dark_bg
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtítulo
        ws.merge_cells('A2:N2')
        ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(color="64B5F6", italic=True, size=12)
        ws['A2'].fill = dark_bg
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Estadísticas
        ws.merge_cells('A4:D4')
        ws['A4'] = "📈 ESTADÍSTICAS"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        ws['A4'].alignment = Alignment(horizontal='center')
        
        # Tarjetas de métricas
        metricas = [
            ("A5", "D5", "🏗️ ACTIVIDADES", "24"),
            ("E5", "H5", "👥 RESPONSABLES", "8"),
            ("I5", "L5", "📅 DÍAS", "90"),
            ("M5", "N5", "✅ COMPLETADO", "45%")
        ]
        
        for start, end, titulo, valor in metricas:
            ws.merge_cells(f'{start}:{end}')
            celda = ws[start]
            celda.value = f"{titulo}\n\n{valor}"
            celda.font = Font(color="FFFFFF", size=11)
            celda.fill = card_fill
            celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Datos de ejemplo
        ws.merge_cells('A8:N8')
        ws['A8'] = "📋 ACTIVIDADES RECIENTES"
        ws['A8'].font = header_font
        ws['A8'].fill = header_fill
        ws['A8'].alignment = Alignment(horizontal='center')
        
        encabezados = ["ID", "Actividad", "Responsable", "Inicio", "Fin", "Estado", "Progreso"]
        for idx, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=9, column=idx, value=encabezado)
            celda.font = Font(color="FFFFFF", bold=True)
            celda.fill = header_fill
            celda.alignment = Alignment(horizontal='center')
        
        actividades = [
            ["PRJ-001", "Planificación Inicial", "Juan Pérez", "2024-01-10", "2024-01-15", "Completado", "100%"],
            ["PRJ-002", "Diseño Arquitectura", "María González", "2024-01-15", "2024-02-05", "En Progreso", "75%"],
            ["PRJ-003", "Desarrollo", "Carlos López", "2024-02-01", "2024-03-15", "Pendiente", "0%"]
        ]
        
        for i, actividad in enumerate(actividades, start=10):
            for j, valor in enumerate(actividad, start=1):
                celda = ws.cell(row=i, column=j, value=valor)
                if j == 6:  # Columna Estado
                    if valor == "Completado":
                        celda.fill = success_fill
                    elif valor == "En Progreso":
                        celda.fill = info_fill
                    elif valor == "Pendiente":
                        celda.fill = warning_fill
        
        # Ajustar anchos de columna
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Crear más hojas
        ws2 = wb.create_sheet("Diagrama Gantt")
        ws2.merge_cells('A1:K1')
        ws2['A1'] = "📅 DIAGRAMA GANTT"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        ws3 = wb.create_sheet("Panel Responsables")
        ws3.merge_cells('A1:J1')
        ws3['A1'] = "👥 PANEL DE RESPONSABLES"
        ws3['A1'].font = title_font
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        # Guardar archivo
        nombre_archivo = f"Mapa_Flujo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_completa = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
        wb.save(ruta_completa)
        
        logger.info(f"Reporte Mapa de Flujo generado: {nombre_archivo}")
        return nombre_archivo, ruta_completa, "Mapa de Flujo generado exitosamente"
        
    except Exception as e:
        logger.error(f"Error generando Mapa de Flujo: {str(e)}")
        raise

# Función para generar Planificación Estratégica
def generar_planificacion():
    """Genera el reporte de Planificación Estratégica"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard Planificación"
        
        # Estilos
        title_font = Font(color="0c2461", bold=True, size=16)
        card_fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
        
        # Título
        ws.merge_cells('A1:O1')
        ws['A1'] = "🎯 PLANIFICACIÓN ESTRATÉGICA"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="e4e8f0", end_color="e4e8f0", fill_type="solid")
        
        # Subtítulo
        ws.merge_cells('A2:O2')
        ws['A2'] = f"Sistema integral de gestión estratégica - {datetime.now().strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(color="64748b", italic=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Métricas
        def crear_metricas(fila, col, titulo, valor):
            ws.merge_cells(start_row=fila, start_column=col, end_row=fila+2, end_column=col+4)
            celda = ws.cell(row=fila, column=col)
            celda.value = f"{titulo}\n\n{valor}"
            celda.font = Font(color="0c2461", bold=True, size=12)
            celda.fill = card_fill
            celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        crear_metricas(4, 1, "ESTRATEGIAS", "15")
        crear_metricas(4, 6, "TÁCTICAS", "45")
        crear_metricas(4, 11, "ACTIVIDADES", "120")
        
        crear_metricas(8, 1, "RESPONSABLES", "25")
        crear_metricas(8, 6, "EJES", "7")
        crear_metricas(8, 11, "PROGRESO", "68%")
        
        # Tabla de estrategias
        ws.merge_cells('A12:F12')
        ws['A12'] = "📋 ESTRATEGIAS ACTIVAS"
        ws['A12'].font = Font(color="0c2461", bold=True, size=14)
        ws['A12'].alignment = Alignment(horizontal='center')
        
        encabezados = ["Eje", "Tipo", "Estrategia", "Tácticas", "Fecha", "Estado"]
        for idx, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=13, column=idx, value=encabezado)
            celda.font = Font(color="ffffff", bold=True)
            celda.fill = PatternFill(start_color="0c2461", end_color="1e3799", fill_type="solid")
            celda.alignment = Alignment(horizontal='center')
        
        estrategias = [
            ["Educación", "FO", "Capacitación continua", "3", "2024-01-15", "Activo"],
            ["Salud", "DO", "Programa bienestar", "2", "2024-01-20", "Activo"],
            ["Empleabilidad", "FA", "Alianzas estratégicas", "4", "2024-01-25", "En progreso"]
        ]
        
        for i, estrategia in enumerate(estrategias, start=14):
            for j, valor in enumerate(estrategia, start=1):
                ws.cell(row=i, column=j, value=valor)
        
        # Ajustar anchos
        for col in range(1, 16):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Más hojas
        ws2 = wb.create_sheet("Estrategias Detalladas")
        ws2.merge_cells('A1:H1')
        ws2['A1'] = "📊 ESTRATEGIAS DETALLADAS"
        ws2['A1'].font = title_font
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        ws3 = wb.create_sheet("Métricas")
        ws3.merge_cells('A1:K1')
        ws3['A1'] = "📈 MÉTRICAS DE SEGUIMIENTO"
        ws3['A1'].font = title_font
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        # Guardar
        nombre_archivo = f"Planificacion_Estrategica_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_completa = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
        wb.save(ruta_completa)
        
        logger.info(f"Reporte Planificación Estratégica generado: {nombre_archivo}")
        return nombre_archivo, ruta_completa, "Planificación Estratégica generada exitosamente"
        
    except Exception as e:
        logger.error(f"Error generando Planificación Estratégica: {str(e)}")
        raise

# Función para generar Análisis FODA
def generar_foda():
    """Genera el reporte de Análisis FODA"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matriz FODA"
        
        # Estilos
        fortalezas_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        debilidades_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        oportunidades_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        amenazas_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        
        # Título
        ws.merge_cells('A1:N1')
        ws['A1'] = "🔄 ANÁLISIS FODA CRUZADO"
        ws['A1'].font = Font(color="2C3E50", bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color="E4E8F0", end_color="E4E8F0", fill_type="solid")
        
        # Subtítulo
        ws.merge_cells('A2:N2')
        ws['A2'] = "Sistema de análisis estratégico combinado"
        ws['A2'].font = Font(color="666666", italic=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Sección Fortalezas
        ws.merge_cells('A4:D4')
        ws['A4'] = "🔰 FORTALEZAS"
        ws['A4'].font = Font(color="FFFFFF", bold=True, size=12)
        ws['A4'].fill = fortalezas_fill
        ws['A4'].alignment = Alignment(horizontal='center')
        
        ws['A5'] = "ID"
        ws['B5'] = "Descripción"
        ws['C5'] = "Prioridad"
        ws['D5'] = "Impacto"
        
        fortalezas = [
            ["F001", "Equipo especializado", "Alta", "Alto"],
            ["F002", "Tecnología actualizada", "Media", "Medio"],
            ["F003", "Marca reconocida", "Alta", "Alto"]
        ]
        
        for i, fortaleza in enumerate(fortalezas, start=6):
            for j, valor in enumerate(fortaleza, start=1):
                ws.cell(row=i, column=j, value=valor)
        
        # Sección Oportunidades
        ws.merge_cells('F4:I4')
        ws['F4'] = "🎯 OPORTUNIDADES"
        ws['F4'].font = Font(color="FFFFFF", bold=True, size=12)
        ws['F4'].fill = oportunidades_fill
        ws['F4'].alignment = Alignment(horizontal='center')
        
        ws['F5'] = "ID"
        ws['G5'] = "Descripción"
        ws['H5'] = "Prioridad"
        ws['I5'] = "Potencial"
        
        oportunidades = [
            ["O001", "Nuevo mercado", "Alta", "Alto"],
            ["O002", "Alianzas posibles", "Media", "Medio"],
            ["O003", "Fondos disponibles", "Alta", "Alto"]
        ]
        
        for i, oportunidad in enumerate(oportunidades, start=6):
            for j, valor in enumerate(oportunidad, start=6):
                ws.cell(row=i, column=j, value=valor)
        
        # Ajustar anchos
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Más hojas
        ws2 = wb.create_sheet("Combinaciones")
        ws2.merge_cells('A1:K1')
        ws2['A1'] = "🔄 COMBINACIONES ESTRATÉGICAS"
        ws2['A1'].font = Font(color="FFFFFF", bold=True, size=14)
        ws2['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        ws3 = wb.create_sheet("Dashboard FODA")
        ws3.merge_cells('A1:G1')
        ws3['A1'] = "📊 DASHBOARD FODA"
        ws3['A1'].font = Font(color="FFFFFF", bold=True, size=14)
        ws3['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        # Guardar
        nombre_archivo = f"Analisis_FODA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_completa = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
        wb.save(ruta_completa)
        
        logger.info(f"Reporte FODA generado: {nombre_archivo}")
        return nombre_archivo, ruta_completa, "Análisis FODA generado exitosamente"
        
    except Exception as e:
        logger.error(f"Error generando FODA: {str(e)}")
        raise

# Función para generar Reporte Completo
def generar_completo():
    """Genera el reporte completo integrado"""
    try:
        wb = openpyxl.Workbook()
        
        # === ÍNDICE ===
        ws_indice = wb.active
        ws_indice.title = "Índice"
        
        ws_indice.merge_cells('A1:E1')
        ws_indice['A1'] = "📚 REPORTE COMPLETO INTEGRADO"
        ws_indice['A1'].font = Font(color="FFFFFF", bold=True, size=18)
        ws_indice['A1'].fill = PatternFill(start_color="0A0F2B", end_color="0A0F2B", fill_type="solid")
        ws_indice['A1'].alignment = Alignment(horizontal='center')
        
        ws_indice.merge_cells('A2:E2')
        ws_indice['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws_indice['A2'].font = Font(color="64B5F6", italic=True)
        ws_indice['A2'].alignment = Alignment(horizontal='center')
        
        # Secciones
        secciones = [
            ("📊 MAPA DE FLUJO", [
                "• Dashboard Resumen",
                "• Diagrama Gantt",
                "• Panel de Responsables"
            ]),
            ("🎯 PLANIFICACIÓN ESTRATÉGICA", [
                "• Estrategias Generadas",
                "• Actividades Detalladas",
                "• Métricas de Seguimiento"
            ]),
            ("🔄 ANÁLISIS FODA", [
                "• Matriz FODA",
                "• Combinaciones",
                "• Dashboard FODA"
            ]),
            ("📈 RESUMEN EJECUTIVO", [
                "• Dashboard Consolidado",
                "• Próximos Pasos",
                "• Recomendaciones"
            ])
        ]
        
        fila_actual = 4
        for titulo, elementos in secciones:
            ws_indice.merge_cells(f'A{fila_actual}:E{fila_actual}')
            ws_indice.cell(row=fila_actual, column=1, value=titulo)
            ws_indice.cell(row=fila_actual, column=1).font = Font(bold=True, size=14, color="0A0F2B")
            fila_actual += 1
            
            for elemento in elementos:
                ws_indice.cell(row=fila_actual, column=2, value=elemento)
                fila_actual += 1
            
            fila_actual += 1
        
        # === DASHBOARD RESUMEN ===
        ws_dashboard = wb.create_sheet("Dashboard Resumen")
        ws_dashboard.merge_cells('A1:N1')
        ws_dashboard['A1'] = "📊 DASHBOARD CONSOLIDADO"
        ws_dashboard['A1'].font = Font(color="FFFFFF", bold=True, size=16)
        ws_dashboard['A1'].fill = PatternFill(start_color="0A0F2B", end_color="0A0F2B", fill_type="solid")
        ws_dashboard['A1'].alignment = Alignment(horizontal='center')
        
        # Métricas
        metricas = [
            ("A3", "D5", "TOTAL REPORTES", "3"),
            ("E3", "H5", "HOJAS", "12"),
            ("I3", "L5", "MÉTRICAS", "24"),
            ("M3", "N5", "ACTUALIZADO", datetime.now().strftime("%d/%m"))
        ]
        
        for start, end, titulo, valor in metricas:
            ws_dashboard.merge_cells(f'{start}:{end}')
            celda = ws_dashboard[start.split('$')[0] + start.split('$')[1][0]]
            celda.value = f"{titulo}\n\n{valor}"
            celda.font = Font(color="FFFFFF", size=11)
            celda.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # === HOJAS ADICIONALES ===
        # Mapa de Flujo
        ws_mapa = wb.create_sheet("Mapa de Flujo")
        ws_mapa['A1'] = "📊 SECCIÓN: MAPA DE FLUJO"
        ws_mapa['A1'].font = Font(bold=True, size=14, color="0A0F2B")
        
        # Planificación
        ws_plan = wb.create_sheet("Planificación")
        ws_plan['A1'] = "🎯 SECCIÓN: PLANIFICACIÓN"
        ws_plan['A1'].font = Font(bold=True, size=14, color="0c2461")
        
        # FODA
        ws_foda = wb.create_sheet("Análisis FODA")
        ws_foda['A1'] = "🔄 SECCIÓN: ANÁLISIS FODA"
        ws_foda['A1'].font = Font(bold=True, size=14, color="2C3E50")
        
        # Resumen Ejecutivo
        ws_resumen = wb.create_sheet("Resumen Ejecutivo")
        ws_resumen.merge_cells('A1:G1')
        ws_resumen['A1'] = "📈 RESUMEN EJECUTIVO"
        ws_resumen['A1'].font = Font(bold=True, size=16, color="0A0F2B")
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        
        contenido = [
            ("Resumen del Reporte:", "Este documento integra todos los análisis estratégicos"),
            ("Fecha de Generación:", datetime.now().strftime("%d/%m/%Y")),
            ("Sistemas Incluidos:", "Mapa de Flujo, Planificación Estratégica, Análisis FODA"),
            ("Total de Hojas:", "12"),
            ("Estado:", "🟢 Completado y listo para revisión"),
            ("Próxima Revisión:", (datetime.now() + timedelta(days=30)).strftime("%d/%m/%Y")),
            ("Responsable:", session.get('user', 'Administrador del Sistema'))
        ]
        
        for i, (titulo, valor) in enumerate(contenido, start=3):
            ws_resumen.cell(row=i, column=1, value=titulo).font = Font(bold=True)
            ws_resumen.cell(row=i, column=3, value=valor)
        
        # Ajustar anchos
        for ws in wb.worksheets:
            for col in range(1, 15):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Eliminar hoja por defecto si existe
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Guardar
        nombre_archivo = f"Reporte_Completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_completa = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
        wb.save(ruta_completa)
        
        logger.info(f"Reporte Completo generado: {nombre_archivo}")
        return nombre_archivo, ruta_completa, "Reporte completo generado exitosamente"
        
    except Exception as e:
        logger.error(f"Error generando Reporte Completo: {str(e)}")
        raise

# Ruta para generar reportes (AJAX)
@app.route('/generar_reporte', methods=['POST'])
@login_required
def generar_reporte():
    """Genera el reporte solicitado"""
    try:
        tipo = request.form.get('tipo')
        usuario = session.get('user', 'Usuario')
        
        logger.info(f"Usuario {usuario} solicitó generar reporte tipo: {tipo}")
        
        if tipo == 'mapa_flujo':
            nombre, ruta, mensaje = generar_mapa_flujo()
        elif tipo == 'planificacion':
            nombre, ruta, mensaje = generar_planificacion()
        elif tipo == 'foda':
            nombre, ruta, mensaje = generar_foda()
        elif tipo == 'completo':
            nombre, ruta, mensaje = generar_completo()
        else:
            return jsonify({
                'success': False,
                'message': 'Tipo de reporte no válido'
            }), 400
        
        # Calcular tamaño
        tamaño = os.path.getsize(ruta)
        
        return jsonify({
            'success': True,
            'message': mensaje,
            'filename': nombre,
            'path': ruta,
            'size': tamaño,
            'download_url': url_for('descargar_reporte', filename=nombre)
        })
        
    except Exception as e:
        logger.error(f"Error en generar_reporte: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al generar el reporte: {str(e)}'
        }), 500

# Ruta para descargar reportes
@app.route('/descargar/<filename>')
@login_required
def descargar_reporte(filename):
    """Descarga un reporte generado"""
    try:
        ruta = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(ruta):
            flash('El archivo solicitado no existe.', 'danger')
            return redirect(url_for('dashboard'))
        
        logger.info(f"Usuario {session.get('user')} descargó: {filename}")
        
        # Registrar descarga en sesión
        if 'descargas' not in session:
            session['descargas'] = []
        session['descargas'].append({
            'archivo': filename,
            'fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
        return send_file(ruta, as_attachment=True)
        
    except Exception as e:
        logger.error(f"Error descargando {filename}: {str(e)}")
        flash(f'Error al descargar el archivo: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

# Ruta para eliminar reportes
@app.route('/eliminar/<filename>', methods=['POST'])
@login_required
def eliminar_reporte(filename):
    """Elimina un reporte generado"""
    try:
        ruta = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(ruta):
            return jsonify({
                'success': False,
                'message': 'El archivo no existe'
            }), 404
        
        os.remove(ruta)
        logger.info(f"Usuario {session.get('user')} eliminó: {filename}")
        
        return jsonify({
            'success': True,
            'message': 'Archivo eliminado exitosamente'
        })
        
    except Exception as e:
        logger.error(f"Error eliminando {filename}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al eliminar el archivo: {str(e)}'
        }), 500

# Ruta para vista previa (simulada)
@app.route('/vista_previa/<filename>')
@login_required
def vista_previa(filename):
    """Muestra información del reporte"""
    ruta = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    if not os.path.exists(ruta):
        flash('El archivo no existe.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Información básica del archivo
    info = {
        'nombre': filename,
        'tamaño': os.path.getsize(ruta),
        'fecha_creacion': datetime.fromtimestamp(os.path.getctime(ruta)),
        'tipo': 'Desconocido'
    }
    
    if 'Mapa_Flujo' in filename:
        info['tipo'] = 'Mapa de Flujo'
        info['icono'] = '📊'
    elif 'Planificacion' in filename:
        info['tipo'] = 'Planificación Estratégica'
        info['icono'] = '🎯'
    elif 'FODA' in filename:
        info['tipo'] = 'Análisis FODA'
        info['icono'] = '🔄'
    elif 'Completo' in filename:
        info['tipo'] = 'Reporte Completo'
        info['icono'] = '📚'
    
    return render_template('vista_previa.html',
                         usuario=session.get('user'),
                         info=info,
                         filename=filename)

# Ruta para configuración
@app.route('/configuracion')
@login_required
def configuracion():
    """Página de configuración del sistema"""
    return render_template('configuracion.html', usuario=session.get('user'))

# Ruta para ayuda
@app.route('/ayuda')
def ayuda():
    """Página de ayuda"""
    return render_template('ayuda.html', usuario=session.get('user', 'Invitado'))

# Ruta para API de estado
@app.route('/api/estado')
def api_estado():
    """API para verificar el estado del sistema"""
    return jsonify({
        'status': 'online',
        'version': '2.0',
        'timestamp': datetime.now().isoformat(),
        'reportes_generados': len(os.listdir(app.config['UPLOAD_FOLDER'])) if os.path.exists(app.config['UPLOAD_FOLDER']) else 0
    })

# Manejo de errores
@app.errorhandler(404)
def pagina_no_encontrada(e):
    return render_template('error.html',
                         error_code=404,
                         error_message="Página no encontrada",
                         usuario=session.get('user', 'Invitado')), 404

@app.errorhandler(500)
def error_servidor(e):
    logger.error(f"Error 500: {str(e)}")
    return render_template('error.html',
                         error_code=500,
                         error_message="Error interno del servidor",
                         usuario=session.get('user', 'Invitado')), 500

@app.errorhandler(413)
def archivo_muy_grande(e):
    return render_template('error.html',
                         error_code=413,
                         error_message="Archivo demasiado grande",
                         usuario=session.get('user', 'Invitado')), 413

# Comando para limpiar reportes antiguos
def limpiar_reportes_antiguos(dias=7):
    """Elimina reportes más antiguos que el número de días especificado"""
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            return 0
        
        limite = datetime.now() - timedelta(days=dias)
        eliminados = 0
        
        for archivo in os.listdir(UPLOAD_FOLDER):
            if archivo.endswith('.xlsx'):
                ruta = os.path.join(UPLOAD_FOLDER, archivo)
                fecha_creacion = datetime.fromtimestamp(os.path.getctime(ruta))
                
                if fecha_creacion < limite:
                    os.remove(ruta)
                    eliminados += 1
                    logger.info(f"Eliminado reporte antiguo: {archivo}")
        
        return eliminados
    except Exception as e:
        logger.error(f"Error limpiando reportes antiguos: {str(e)}")
        return 0

# Punto de entrada principal
if __name__ == '__main__':
    # Crear carpetas necesarias
    if not os.path.exists('templates'):
        os.makedirs('templates')
    
    if not os.path.exists('static'):
        os.makedirs('static')
        os.makedirs('static/css')
        os.makedirs('static/js')
        os.makedirs('static/img')
    
    # Generar templates básicos si no existen
    templates_necesarios = {
        'index.html': '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sistema de Reportes Excel</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.8.1/font/bootstrap-icons.css">
    <style>
        .hero-section {
            background: linear-gradient(135deg, #0A0F2B 0%, #1A237E 100%);
            color: white;
            padding: 100px 0;
        }
        .feature-card {
            border: none;
            border-radius: 15px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.1);
            transition: transform 0.3s;
            height: 100%;
        }
        .feature-card:hover {
            transform: translateY(-5px);
        }
        .card-icon {
            font-size: 3rem;
            margin-bottom: 20px;
        }
        .stat-card {
            background: white;
            border-radius: 10px;
            padding: 20px;
            box-shadow: 0 3px 10px rgba(0,0,0,0.1);
        }
    </style>
</head>
<body>
    <!-- Navbar -->
    <nav class="navbar navbar-expand-lg navbar-dark" style="background-color: #0A0F2B;">
        <div class="container">
            <a class="navbar-brand" href="/">
                <i class="bi bi-file-earmark-spreadsheet"></i>
                Sistema Reportes
            </a>
            <div class="navbar-nav ms-auto">
                {% if usuario == 'Invitado' %}
                <a class="nav-link" href="/login">
                    <i class="bi bi-box-arrow-in-right"></i> Iniciar Sesión
                </a>
                {% else %}
                <span class="navbar-text me-3">
                    <i class="bi bi-person-circle"></i> {{ usuario }}
                </span>
                <a class="nav-link" href="/dashboard">
                    <i class="bi bi-speedometer2"></i> Dashboard
                </a>
                <a class="nav-link" href="/logout">
                    <i class="bi bi-box-arrow-right"></i> Salir
                </a>
                {% endif %}
            </div>
        </div>
    </nav>

    <!-- Hero Section -->
    <section class="hero-section text-center">
        <div class="container">
            <h1 class="display-4 mb-4">
                <i class="bi bi-file-earmark-spreadsheet"></i>
                Sistema de Reportes Excel
            </h1>
            <p class="lead mb-4">
                Genera reportes profesionales de gestión, planificación y análisis estratégico
            </p>
            {% if usuario == 'Invitado' %}
            <a href="/login" class="btn btn-light btn-lg me-3">
                <i class="bi bi-box-arrow-in-right"></i> Comenzar
            </a>
            {% else %}
            <a href="/dashboard" class="btn btn-light btn-lg me-3">
                <i class="bi bi-speedometer2"></i> Ir al Dashboard
            </a>
            <a href="/generar" class="btn btn-outline-light btn-lg">
                <i class="bi bi-plus-circle"></i> Nuevo Reporte
            </a>
            {% endif %}
        </div>
    </section>

    <!-- Features -->
    <section class="py-5">
        <div class="container">
            <h2 class="text-center mb-5">Tipos de Reportes Disponibles</h2>
            <div class="row g-4">
                <div class="col-md-3">
                    <div class="card feature-card text-center p-4">
                        <div class="card-icon text-primary">
                            <i class="bi bi-kanban"></i>
                        </div>
                        <h4>Mapa de Flujo</h4>
                        <p>Gestión de actividades y responsabilidades con diagramas Gantt</p>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card feature-card text-center p-4">
                        <div class="card-icon text-success">
                            <i class="bi bi-bullseye"></i>
                        </div>
                        <h4>Planificación</h4>
                        <p>Estrategias, tácticas y actividades organizadas</p>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card feature-card text-center p-4">
                        <div class="card-icon text-warning">
                            <i class="bi bi-arrow-repeat"></i>
                        </div>
                        <h4>Análisis FODA</h4>
                        <p>Matriz estratégica con combinaciones múltiples</p>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="card feature-card text-center p-4">
                        <div class="card-icon text-info">
                            <i class="bi bi-collection"></i>
                        </div>
                        <h4>Reporte Completo</h4>
                        <p>Todos los sistemas integrados en un solo documento</p>
                    </div>
                </div>
            </div>
        </div>
    </section>

    <!-- Stats -->
    <section class="py-5 bg-light">
        <div class="container">
            <div class="row text-center">
                <div class="col-md-3">
                    <div class="stat-card">
                        <h3 class="text-primary">20+</h3>
                        <p>Formatos de Reporte</p>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="stat-card">
                        <h3 class="text-success">100%</h3>
                        <p>Personalizable</p>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="stat-card">
                        <h3 class="text-warning">24/7</h3>
                        <p>Disponibilidad</p>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="stat-card">
                        <h3 class="text-info">0</h3>
                        <p>Requerimientos Instalación</p>
                    </div>
                </div>
            </div>
        </div>
    </section>

    <!-- Footer -->
    <footer class="bg-dark text-white py-4">
        <div class="container text-center">
            <p>&copy; 2024 Sistema de Reportes Excel. Todos los derechos reservados.</p>
            <div class="mt-3">
                <a href="/ayuda" class="text-white me-3">
                    <i class="bi bi-question-circle"></i> Ayuda
                </a>
                <a href="/configuracion" class="text-white me-3">
                    <i class="bi bi-gear"></i> Configuración
                </a>
                <a href="/api/estado" class="text-white">
                    <i class="bi bi-heart-pulse"></i> Estado del Sistema
                </a>
            </div>
        </div>
    </footer>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>''',
        
        'login.html': '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Iniciar Sesión - Sistema de Reportes</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.8.1/font/bootstrap-icons.css">
    <style>
        body {
            background: linear-gradient(135deg, #0A0F2B 0%, #1A237E 100%);
            height: 100vh;
            display: flex;
            align-items: center;
        }
        .login-card {
            background: white;
            border-radius: 20px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        .login-header {
            background: #0A0F2B;
            color: white;
            padding: 30px;
            text-align: center;
        }
        .login-body {
            padding: 40px;
        }
        .form-control {
            border-radius: 10px;
            padding: 12px;
            border: 2px solid #e0e0e0;
        }
        .form-control:focus {
            border-color: #1A237E;
            box-shadow: none;
        }
        .btn-login {
            background: #1A237E;
            color: white;
            padding: 12px;
            border-radius: 10px;
            font-weight: bold;
            width: 100%;
        }
        .btn-login:hover {
            background: #0A0F2B;
            color: white;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="row justify-content-center">
            <div class="col-md-5">
                <div class="login-card">
                    <div class="login-header">
                        <h2>
                            <i class="bi bi-file-earmark-spreadsheet"></i><br>
                            Sistema de Reportes
                        </h2>
                        <p class="mb-0">Inicie sesión para acceder al sistema</p>
                    </div>
                    
                    <div class="login-body">
                        {% with messages = get_flashed_messages(with_categories=true) %}
                            {% if messages %}
                                {% for category, message in messages %}
                                    <div class="alert alert-{{ category }} alert-dismissible fade show">
                                        {{ message }}
                                        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
                                    </div>
                                {% endfor %}
                            {% endif %}
                        {% endwith %}
                        
                        <form method="POST" action="/login">
                            <div class="mb-4">
                                <label class="form-label">Usuario</label>
                                <input type="text" name="username" class="form-control" 
                                       placeholder="Ingrese su usuario" required>
                            </div>
                            
                            <div class="mb-4">
                                <label class="form-label">Contraseña</label>
                                <input type="password" name="password" class="form-control" 
                                       placeholder="Ingrese su contraseña" required>
                            </div>
                            
                            <button type="submit" class="btn btn-login mb-3">
                                <i class="bi bi-box-arrow-in-right"></i> Iniciar Sesión
                            </button>
                            
                            <div class="text-center">
                                <a href="/" class="text-decoration-none">
                                    <i class="bi bi-arrow-left"></i> Volver al inicio
                                </a>
                            </div>
                        </form>
                        
                        <div class="alert alert-info mt-4">
                            <small>
                                <i class="bi bi-info-circle"></i>
                                <strong>Credenciales de prueba:</strong><br>
                                Usuario: <code>admin</code><br>
                                Contraseña: <code>admin123</code>
                            </small>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>''',
        
        'dashboard.html': '''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Dashboard - Sistema de Reportes</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.8.1/font/bootstrap-icons.css">
    <style>
        .sidebar {
            background: #0A0F2B;
            color: white;
            min-height: 100vh;
            padding: 0;
        }
        .sidebar-brand {
            padding: 20px;
            text-align: center;
            border-bottom: 1px solid rgba(255,255,255,0.1);
        }
        .nav-link {
            color: rgba(255,255,255,0.8);
            padding: 15px 20px;
            border-left: 4px solid transparent;
        }
        .nav-link:hover {
            color: white;
            background: rgba(255,255,255,0.1);
            border-left-color: #2196F3;
        }
        .nav-link.active {
            color: white;
            background: rgba(255,255,255,0.1);
            border-left-color: #2196F3;
        }
        .stat-card {
            background: white;
            border-radius: 10px;
            padding: 20px;
            box-shadow: 0 3px 10px rgba(0,0,0,0.1);
            border-top: 4px solid;
        }
        .stat-card.total { border-color: #0A0F2B; }
        .stat-card.mapa { border-color: #2196F3; }
        .stat-card.plan { border-color: #4CAF50; }
        .stat-card.foda { border-color: #FF9800; }
        .stat-card.completo { border-color: #9C27B0; }
        .table-reportes {
            background: white;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 3px 10px rgba(0,0,0,0.1);
        }
        .btn-generar {
            background: linear-gradient(135deg, #1A237E 0%, #0A0F2B 100%);
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 8px;
            font-weight: bold;
        }
        .btn-generar:hover {
            color: white;
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(26, 35, 126, 0.3);
        }
    </style>
</head>
<body>
    <div class="container-fluid">
        <div class="row">
            <!-- Sidebar -->
            <div class="col-md-2 sidebar">
                <div class="sidebar-brand">
                    <h4>
                        <i class="bi bi-file-earmark-spreadsheet"></i><br>
                        Reportes Excel
                    </h4>
                    <small class="text-muted">{{ usuario }}</small>
                </div>
                
                <nav class="nav flex-column mt-3">
                    <a class="nav-link active" href="/dashboard">
                        <i class="bi bi-speedometer2"></i> Dashboard
                    </a>
                    <a class="nav-link" href="/generar">
                        <i class="bi bi-plus-circle"></i> Generar Reporte
                    </a>
                    <a class="nav-link" href="#">
                        <i class="bi bi-folder"></i> Mis Reportes
                    </a>
                    <a class="nav-link" href="/configuracion">
                        <i class="bi bi-gear"></i> Configuración
                    </a>
                    <a class="nav-link" href="/ayuda">
                        <i class="bi bi-question-circle"></i> Ayuda
                    </a>
                    <a class="nav-link" href="/logout">
                        <i class="bi bi-box-arrow-right"></i> Salir
                    </a>
                </nav>
            </div>
            
            <!-- Main Content -->
            <div class="col-md-10 p-4">
                <!-- Header -->
                <div class="d-flex justify-content-between align-items-center mb-4">
                    <h2>
                        <i class="bi bi-speedometer2"></i>
                        Dashboard Principal
                    </h2>
                    <a href="/generar" class="btn btn-generar">
                        <i class="bi bi-plus-circle"></i> Nuevo Reporte
                    </a>
                </div>
                
                <!-- Stats Cards -->
                <div class="row mb-4">
                    <div class="col-md-2">
                        <div class="stat-card total">
                            <h5 class="text-muted">Total</h5>
                            <h3>{{ estadisticas.total_reportes }}</h3>
                            <small>Reportes</small>
                        </div>
                    </div>
                    <div class="col-md-2">
                        <div class="stat-card mapa">
                            <h5 class="text-primary">Mapa Flujo</h5>
                            <h3>{{ estadisticas.total_mapa_flujo }}</h3>
                        </div>
                    </div>
                    <div class="col-md-2">
                        <div class="stat-card plan">
                            <h5 class="text-success">Planificación</h5>
                            <h3>{{ estadisticas.total_planificacion }}</h3>
                        </div>
                    </div>
                    <div class="col-md-2">
                        <div class="stat-card foda">
                            <h5 class="text-warning">FODA</h5>
                            <h3>{{ estadisticas.total_foda }}</h3>
                        </div>
                    </div>
                    <div class="col-md-2">
                        <div class="stat-card completo">
                            <h5 class="text-info">Completos</h5>
                            <h3>{{ estadisticas.total_completo }}</h3>
                        </div>
                    </div>
                    <div class="col-md-2">
                        <div class="stat-card">
                            <h5 class="text-muted">Espacio</h5>
                            <h3>{{ "%.1f"|format(estadisticas.tamaño_total / (1024*1024)) }} MB</h3>
                        </div>
                    </div>
                </div>
                
                <!-- Reportes Recientes -->
                <div class="card table-reportes">
                    <div class="card-header bg-primary text-white">
                        <h5 class="mb-0">
                            <i class="bi bi-clock-history"></i>
                            Reportes Recientes
                        </h5>
                    </div>
                    <div class="card-body p-0">
                        {% if reportes %}
                        <div class="table-responsive">
                            <table class="table table-hover mb-0">
                                <thead class="table-light">
                                    <tr>
                                        <th>Nombre</th>
                                        <th>Tamaño</th>
                                        <th>Fecha</th>
                                        <th>Acciones</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    {% for reporte in reportes %}
                                    <tr>
                                        <td>
                                            {% if 'Mapa_Flujo' in reporte.nombre %}
                                            <i class="bi bi-kanban text-primary"></i>
                                            {% elif 'Planificacion' in reporte.nombre %}
                                            <i class="bi bi-bullseye text-success"></i>
                                            {% elif 'FODA' in reporte.nombre %}
                                            <i class="bi bi-arrow-repeat text-warning"></i>
                                            {% elif 'Completo' in reporte.nombre %}
                                            <i class="bi bi-collection text-info"></i>
                                            {% endif %}
                                            {{ reporte.nombre }}
                                        </td>
                                        <td>{{ "%.1f"|format(reporte.tamaño / 1024) }} KB</td>
                                        <td>{{ reporte.fecha }}</td>
                                        <td>
                                            <a href="{{ url_for('descargar_reporte', filename=reporte.nombre) }}" 
                                               class="btn btn-sm btn-success" title="Descargar">
                                                <i class="bi bi-download"></i>
                                            </a>
                                            <a href="{{ url_for('vista_previa', filename=reporte.nombre) }}" 
                                               class="btn btn-sm btn-info" title="Vista Previa">
                                                <i class="bi bi-eye"></i>
                                            </a>
                                            <button class="btn btn-sm btn-danger" 
                                                    onclick="eliminarReporte('{{ reporte.nombre }}')"
                                                    title="Eliminar">
                                                <i class="bi bi-trash"></i>
                                            </button>
                                        </td>
                                    </tr>
                                    {% endfor %}
                                </tbody>
                            </table>
                        </div>
                        {% else %}
                        <div class="text-center py-5">
                            <i class="bi bi-folder-x" style="font-size: 3rem; color: #ccc;"></i>
                            <h5 class="mt-3 text-muted">No hay reportes generados</h5>
                            <a href="/generar" class="btn btn-primary mt-2">
                                <i class="bi bi-plus-circle"></i> Crear Primer Reporte
                            </a>
                        </div>
                        {% endif %}
                    </div>
                </div>
                
                <!-- Quick Actions -->
                <div class="row mt-4">
                    <div class="col-md-4">
                        <div class="card">
                            <div class="card-body text-center">
                                <i class="bi bi-kanban" style="font-size: 3rem; color: #2196F3;"></i>
                                <h5 class="mt-3">Mapa de Flujo</h5>
                                <p class="text-muted">Gestión de proyectos</p>
                                <a href="/generar?tipo=mapa_flujo" class="btn btn-outline-primary">
                                    Generar
                                </a>
                            </div>
                        </div>
                    </div>
                    <div class="col-md-4">
                        <div class="card">
                            <div class="card-body text-center">
                                <i class="bi bi-bullseye" style="font-size: 3rem; color: #4CAF50;"></i>
                                <h5 class="mt-3">Planificación</h5>
                                <p class="text-muted">Estrategias y tácticas</p>
                                <a href="/generar?tipo=planificacion" class="btn btn-outline-success">
                                    Generar
                                </a>
                            </div>
                        </div>
                    </div>
                    <div class="col-md-4">
                        <div class="card">
                            <div class="card-body text-center">
                                <i class="bi bi-arrow-repeat" style="font-size: 3rem; color: #FF9800;"></i>
                                <h5 class="mt-3">Reporte Completo</h5>
                                <p class="text-muted">Todos los sistemas</p>
                                <a href="/generar?tipo=completo" class="btn btn-outline-warning">
                                    Generar
                                </a>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        function eliminarReporte(filename) {
            if (confirm('¿Está seguro de eliminar este reporte?')) {
                fetch('/eliminar/' + filename, {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    }
                })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        alert('Reporte eliminado exitosamente');
                        location.reload();
                    } else {
                        alert('Error: ' + data.message);
                    }
                })
                .catch(error => {
                    alert('Error al eliminar el reporte');
                });
            }
        }
    </script>
</body>
</html>'''
    }
    
    # Crear templates
    for nombre, contenido in templates_necesarios.items():
        ruta_template = os.path.join('templates', nombre)
        if not os.path.exists(ruta_template):
            with open(ruta_template, 'w', encoding='utf-8') as f:
                f.write(contenido)
            print(f"✅ Template creado: {nombre}")
    
    # Limpiar reportes antiguos al inicio
    eliminados = limpiar_reportes_antiguos(30)
    if eliminados > 0:
        logger.info(f"Eliminados {eliminados} reportes antiguos")
    
    # Iniciar la aplicación
    print("\n" + "="*60)
    print("🚀 SISTEMA DE REPORTES EXCEL - APLICACIÓN WEB")
    print("="*60)
    print(f"\n📁 Carpeta de reportes: {UPLOAD_FOLDER}")
    print(f"🔗 URL de acceso: http://localhost:5000")
    print(f"🔐 Usuario de prueba: admin / admin123")
    print("\n📌 Comandos útiles:")
    print("   • Ctrl+C para detener el servidor")
    print("   • Recargar página para actualizar reportes")
    print("\n✅ Servidor iniciado correctamente")
    print("="*60 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=True)
